from PySide2.QtWidgets import QApplication, QFileDialog, QMessageBox
from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QFile  #这段要补上，笔记文件里没有
import xlrd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.styles import Font

class PyEx:

    def __init__(self):
        # 从文件中加载UI定义，下面要补上，笔记文件里没有
        qfile_stats = QFile("ui/PyEx.ui") #加载的ui文件
        qfile_stats.open(QFile.ReadOnly)#这句和下句是固定写法
        qfile_stats.close()#关闭

        # 从 UI 定义中动态 创建一个相应的窗口对象
        # 注意：里面的控件对象也成为窗口对象的属性了
        # 比如 self.ui.button , self.ui.textEdit
        self.ui = QUiLoader().load('ui/PyEx.ui')

        self.ui.pushButton.clicked.connect(self.fileLoader)
        self.ui.saveButton.clicked.connect(self.saveButton)
        self.ui.plotButton.clicked.connect(self.autuPlot)

    #导入按钮
    def fileLoader(self):
        self.filePath,type = QFileDialog.getOpenFileName(
            self.ui,  # 父窗口对象
            "选择你要上传的excel",  # 标题
            r"c:\\",  # 起始目录
            "文件类型 (*.xls, *.xlsx)"  # 选择类型过滤项，过滤内容在括号中
        )
        self.ui.textBrowser.clear()
        self.ui.textBrowser.append(self.filePath)

    #导出按钮
    def saveButton(self):
        try:
            saveFilePath = QFileDialog.getExistingDirectory(self.ui, "选择存储路径")
            number = self.ui.spinBox.value() #获取数字，int类型
            text = self.ui.KeyLineEdit.text()#读取关键词文本框内容
            #读取Excel数据
            book = xlrd.open_workbook(self.filePath)
            sheet = book.sheet_by_index(0)

            sheet_col = sheet.col_values(colx=number)
            sheet_row = sheet.row_values(rowx=0)#表头
            #检索有指定有该关键词的行
            liRowAccess = []#记录检索通过的数据的行数索引
            liAccess = []#记录检索通过的数据的行数索引，与上一个列表索引相同，一一对应。
            i = 0
            for row, content in enumerate(sheet_col): #真实row要加1]
                if row>0 and type(content) is str and (text in content):
                    liRowAccess.append(row)
                    liAccess.append(content)
                else:
                    continue
                i+=1

            #过滤后的每行存入二维数组中，【行坐标，0开始，到最后一个数据结束】【列坐标，0开始到最后一个标题结束】
            filter = [[' ' for i in range(len(sheet_row))] for i in range(len(liAccess))]#二维数组初始化
            for r in range(len(liAccess)):
                for c in range(len(sheet_row)):
                    filter[r][c] = sheet.row_values(rowx=liRowAccess[r])[c]

            #新建excel
            # 创建一个Excel workbook 对象
            book = openpyxl.Workbook()
            # 创建时，会自动产生一个sheet，通过active获取
            sh = book.active
            formName = self.ui.KeyLineEdit_2.text()#获取表名
            fileName = self.ui.FileLineEdit.text() # 获取文件名
            sh.title = formName

            #写入标题栏
            for i in range(len(sheet_row)):
                sh.cell(1, i+1).value = sheet_row[i]#openpyxl库中行号列号从1开始
                sh.cell(1, i+1).font = Font(
                        size=12,    # 设定文字大小
                        bold=True,  # 设定为粗体
                        )


            #写入数据
            for r in range(len(liAccess)):
                for c in range(len(sheet_row)):
                    sh.cell(r+2,c+1).value = filter[r][c]

            book.save(saveFilePath+'/'+fileName+'.xlsx')
            #导出成功提示
            QMessageBox.information(
                self.ui,
                '导出成功',
                '导出Excel成功，请到您指定的路径下查看')
        except:
            #给个弹出框显示无导入文件，请先导入文件。
            QMessageBox.critical(
                self.ui,
                '错误',
                '导入文件错误！')

    #自动绘图
    def autuPlot(self):
        try:
            #读取QT中各对象的内容
            titleColNum = self.ui.spinBox_title.value()
            dataColNum = self.ui.spinBox_data.value()
            plotType = self.ui.comboBox.currentText()

            #读取路径下两列数据成一个列表
            book = xlrd.open_workbook(self.filePath)
            sheet = book.sheet_by_index(0)
            titleColList = sheet.col_values(colx=titleColNum)
            dataColList = sheet.col_values(colx=dataColNum)#注意[0]为标题项

            plt.rcParams['font.family'] = 'sans-serif'
            # 设定字体为微软雅黑
            plt.rcParams['font.sans-serif'] = ['Microsoft Yahei']

            if plotType == "折线图":
                plt.plot(titleColList[1::], dataColList[1::])
                plt.xlabel(titleColList[0])
                plt.ylabel(dataColList[0])
                plt.show()
            else:
                plt.bar(titleColList[1::], dataColList[1::])
                plt.xlabel(titleColList[0])
                plt.ylabel(dataColList[0])
                plt.show()
                #绘图成功提示

        except:
            #给个弹出框显示无导入文件，请先导入文件。
            QMessageBox.critical(
                self.ui,
                '错误',
                '读取导入文件错误！')

app = QApplication([])
PyEx = PyEx()
PyEx.ui.show()
app.exec_()